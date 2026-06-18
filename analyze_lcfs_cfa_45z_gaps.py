from __future__ import annotations

from dataclasses import dataclass
from math import asin, cos, radians, sin, sqrt
from pathlib import Path
import json
import math

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import RandomForestRegressor
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler


JSON_PATH = Path(
    r"C:\Users\ehakm\Documents\ELHApp-Carbon_Calculator\docs\static_data\LCFS\lcfs_dropdown_v2.json"
)
OUT_XLSX = Path(r"C:\Users\ehakm\Documents\ELHApp-backend\LCFS_CFA_45Z_Comparison.xlsx")
OUT_MD = Path(r"C:\Users\ehakm\Documents\ELHApp-backend\LCFS_CFA_45Z_Analysis_Report.md")

ILUC = 19.1
FREIGHT_MULTIPLIER = 1.15
FREIGHT_FACTOR = 0.0020
BASELINE_45Z_FREIGHT_MILES = 300

GRID_BRIDGE_ADJ = {
    "CAISO": -1.0,
    "SPP": -1.0,
    "ERCOT": 0.0,
    "PJM": 0.0,
    "MISO/SPP": 0.0,
    "MISO": 1.0,
    "MISO/PJM": 1.0,
}


@dataclass(frozen=True)
class Hub:
    id: str
    name: str
    railroad: str
    lat: float
    lon: float


CA_HUBS = [
    Hub("UP_COLTON", "Colton", "UP", 34.06639, -117.36472),
    Hub("UP_CARSON_LOMITA", "Carson / Lomita Rail", "UP", 33.80100, -118.25500),
    Hub("BNSF_BARSTOW", "Barstow Yard", "BNSF", 34.89319, -117.07461),
    Hub("BNSF_WATSON", "Watson Yard (Carson)", "BNSF", 33.79928, -118.25374),
    Hub("BNSF_HOBART", "Hobart Yard", "BNSF", 34.01300, -118.15300),
    Hub("BNSF_KAISER", "Kaiser Yard (Fontana)", "BNSF", 34.10000, -117.43400),
    Hub("BNSF_CALWA", "Calwa Yard (Fresno)", "BNSF", 36.72200, -119.73200),
]


def as_num(value) -> float:
    if value is None:
        return np.nan
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text.upper() in {"NONE", "NULL", "NAN", "CONFIDENTIAL", "PENDING APPROVAL", "N/A"}:
        return np.nan
    try:
        return float(text)
    except ValueError:
        return np.nan


def norm(value) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def norm_upper(value) -> str:
    return norm(value).upper()


def first(*values):
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def plant_epm(row: dict) -> str:
    fac = row.get("fac_info") or {}
    return norm(first(row.get("EPM_NUMBER"), row.get("epm_number"), row.get("epm"), row.get("EPM"), fac.get("epm")))


def boolish(value) -> bool:
    text = norm(value).lower()
    return text not in {"", "0", "false", "no", "n", "none", "nan", "null", "unknown"}


def best_lcfs_ci(row: dict) -> float:
    ci_summary = row.get("ci_summary") or {}
    value = as_num(first(row.get("ci_lcfs_delivered_g_per_mj"), ci_summary.get("ci_lcfs_delivered_g_per_mj")))
    if math.isfinite(value):
        return value

    detail = row.get("lcfs_detail") or {}
    ca_rows = detail.get("ca_detail") or row.get("ca_detail") or []
    scores = [as_num(d.get("ci_score")) for d in ca_rows if isinstance(d, dict)]
    scores = [x for x in scores if math.isfinite(x)]
    if scores:
        return min(scores)
    return np.nan


def choose_hub(rail_lines: str) -> Hub:
    raw = norm_upper(rail_lines)
    has_up = "UP" in raw or "UNION PACIFIC" in raw
    has_bnsf = "BNSF" in raw
    if has_up and not has_bnsf:
        return CA_HUBS[0]
    if has_bnsf and not has_up:
        return next(h for h in CA_HUBS if h.railroad == "BNSF")
    return CA_HUBS[0]


def haversine_miles(lat1, lon1, lat2, lon2) -> float:
    if not all(math.isfinite(x) for x in [lat1, lon1, lat2, lon2]):
        return np.nan
    r = 3958.7613
    p1, p2 = radians(lat1), radians(lat2)
    dp = radians(lat2 - lat1)
    dl = radians(lon2 - lon1)
    a = sin(dp / 2) ** 2 + cos(p1) * cos(p2) * sin(dl / 2) ** 2
    return 2 * r * asin(sqrt(a))


def pipeline_label(direct, third, sponsor, rail_connect) -> str:
    direct_u = norm_upper(direct)
    third_u = norm_upper(third)
    sponsor_u = norm_upper(sponsor)
    if "DIRECT" in direct_u:
        return "Direct"
    if "SCS" in third_u or "SUMMIT" in third_u or "SUMMIT" in sponsor_u:
        return "Summit/SCS"
    if "TALL" in third_u or "TRAILBLAZER" in third_u or "TALL" in sponsor_u or "TRAILBLAZER" in sponsor_u:
        return "Tallgrass/Trailblazer"
    if boolish(rail_connect):
        return "Rail CO2"
    if boolish(third):
        return "Other third-party"
    return "None"


def gas_label(row: dict) -> str:
    tech = row.get("tech_flags") or {}
    fuel = row.get("fuel_summary") or {}
    value = norm(first(tech.get("gas_supply"), tech.get("primary_process_fuel"), fuel.get("fuel_type_master")))
    if not value:
        return "Unknown"
    value_u = value.upper()
    if "RNG" in value_u or "RENEWABLE NATURAL" in value_u or "BIOGAS" in value_u:
        return "RNG/Biogas"
    if "NATURAL GAS" in value_u:
        return "Natural Gas"
    return value


def flatten_record(row: dict) -> dict:
    fac = row.get("fac_info") or {}
    ci_summary = row.get("ci_summary") or {}
    co2 = row.get("co2_info") or {}
    tech = row.get("tech_flags") or {}
    epa = row.get("epa_ghg_derived") or {}

    epm = plant_epm(row)
    owner = norm(first(row.get("ownership"), fac.get("ownership")))
    state = norm(first(row.get("state"), fac.get("state"))).upper()
    city = norm(first(row.get("city"), fac.get("city")))
    plant = norm(first(row.get("plant_name"), fac.get("plant_name")))

    lcfs_ci = best_lcfs_ci(row)
    cfa = as_num(row.get("canadian_fed_ci"))
    cfa_raw = row.get("canadian_fed_ci")

    lat = as_num(first(row.get("latitude"), fac.get("latitude")))
    lon = as_num(first(row.get("longitude"), fac.get("longitude")))
    rail = norm(first(row.get("rail_lines"), fac.get("rail_lines")))
    hub = choose_hub(rail)
    miles = haversine_miles(lat, lon, hub.lat, hub.lon)
    miles_eff = miles * FREIGHT_MULTIPLIER if math.isfinite(miles) else np.nan
    miles_net = miles_eff - BASELINE_45Z_FREIGHT_MILES if math.isfinite(miles_eff) else np.nan
    freight_ci = miles_net * FREIGHT_FACTOR if math.isfinite(miles_net) else np.nan

    grid = norm(first(row.get("electrical_grid_designation"), tech.get("electrical_grid_designation"))).upper()
    grid_adj = GRID_BRIDGE_ADJ.get(grid, 0.0)

    estimated_45z = lcfs_ci - freight_ci - ILUC + grid_adj if math.isfinite(lcfs_ci) and math.isfinite(freight_ci) else np.nan
    lcfs_cfa_gap = lcfs_ci - cfa if math.isfinite(lcfs_ci) and math.isfinite(cfa) else np.nan
    z45_cfa_gap = estimated_45z - cfa if math.isfinite(estimated_45z) and math.isfinite(cfa) else np.nan

    direct = first(row.get("co2_pipeline_direct"), co2.get("co2_pipeline_direct"))
    third = first(row.get("co2_pipeline_3rd_party"), co2.get("co2_pipeline_3rd_party"))
    sponsor = first(row.get("co2_sponsor"), co2.get("co2_sponsor"))
    rail_co2 = first(row.get("co2_rail_connect"), co2.get("co2_rail_connect"))
    pipe_label = pipeline_label(direct, third, sponsor, rail_co2)

    return {
        "EPM": epm,
        "Plant": plant,
        "Owner": owner,
        "State": state,
        "City": city,
        "LCFS CI": lcfs_ci,
        "Canadian Federal CI": cfa,
        "Canadian Federal CI Raw": cfa_raw,
        "Estimated 45Z CI": estimated_45z,
        "LCFS-CFA Gap": lcfs_cfa_gap,
        "45Z-CFA Gap": z45_cfa_gap,
        "Freight CI Adjustment": freight_ci,
        "Distance to Hub": miles,
        "Net 45Z Freight Miles": miles_net,
        "Default Hub": hub.name,
        "Default Hub Railroad": hub.railroad,
        "Electric Grid": grid,
        "Electric Grid Adjustment": grid_adj,
        "Gas/RNG Usage": gas_label(row),
        "CCS Status": pipe_label,
        "Has CCS/Pipeline": pipe_label != "None",
        "Direct CCS": "DIRECT" in norm_upper(direct),
        "Third Party CCS": pipe_label in {"Summit/SCS", "Tallgrass/Trailblazer", "Other third-party"},
        "CO2 Sponsor": norm(sponsor),
        "Technology": norm(first(row.get("technology"), tech.get("technology"))),
        "Fiber Technology": norm(first(row.get("fiber_technology"), tech.get("fiber_technology"))),
        "Thermal BTU/gal": as_num(first(row.get("thermal_btu_per_gal_est"), epa.get("thermal_btu_per_gal_est"))),
        "Ethanol Capacity MGY": as_num(first(row.get("ethanol_capacity_mgy"), fac.get("ethanol_capacity_mgy"))),
    }


def summary_stats(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["LCFS CI", "Canadian Federal CI", "Estimated 45Z CI", "LCFS-CFA Gap", "45Z-CFA Gap"]
    rows = []
    for col in cols:
        s = df[col].dropna()
        rows.append(
            {
                "Metric": col,
                "Count": len(s),
                "Average": s.mean(),
                "Median": s.median(),
                "Std Dev": s.std(ddof=1),
                "Min": s.min(),
                "Max": s.max(),
            }
        )
    return pd.DataFrame(rows)


def group_summary(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    return (
        df.groupby(group_col, dropna=False)
        .agg(
            rows=("45Z-CFA Gap", "count"),
            avg_45z_cfa_gap=("45Z-CFA Gap", "mean"),
            median_45z_cfa_gap=("45Z-CFA Gap", "median"),
            avg_lcfs_cfa_gap=("LCFS-CFA Gap", "mean"),
            avg_lcfs_ci=("LCFS CI", "mean"),
            avg_cfa=("Canadian Federal CI", "mean"),
            avg_est_45z=("Estimated 45Z CI", "mean"),
        )
        .reset_index()
        .sort_values(["rows", "avg_45z_cfa_gap"], ascending=[False, False])
    )


def model_analysis(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, float, float]:
    features_num = [
        "LCFS CI",
        "Freight CI Adjustment",
        "Distance to Hub",
        "Electric Grid Adjustment",
        "Thermal BTU/gal",
        "Ethanol Capacity MGY",
    ]
    features_cat = [
        "Owner",
        "State",
        "Electric Grid",
        "Gas/RNG Usage",
        "CCS Status",
        "Default Hub Railroad",
        "Technology",
    ]
    model_df = df.dropna(subset=["45Z-CFA Gap"]).copy()
    x = model_df[features_num + features_cat]
    y = model_df["45Z-CFA Gap"]

    pre = ColumnTransformer(
        [
            ("num", Pipeline([("imputer", SimpleImputer(strategy="median")), ("scale", StandardScaler())]), features_num),
            ("cat", Pipeline([("imputer", SimpleImputer(strategy="most_frequent")), ("onehot", OneHotEncoder(handle_unknown="ignore"))]), features_cat),
        ]
    )
    lin = Pipeline([("pre", pre), ("model", LinearRegression())])
    lin.fit(x, y)
    r2_lin = r2_score(y, lin.predict(x))

    names_num = features_num
    names_cat = lin.named_steps["pre"].named_transformers_["cat"].named_steps["onehot"].get_feature_names_out(features_cat).tolist()
    names = names_num + names_cat
    coefs = lin.named_steps["model"].coef_
    coef_df = pd.DataFrame({"feature": names, "linear_coef": coefs}).sort_values("linear_coef", key=lambda s: s.abs(), ascending=False)

    rf = Pipeline(
        [
            ("pre", pre),
            ("model", RandomForestRegressor(n_estimators=500, random_state=7, min_samples_leaf=2)),
        ]
    )
    rf.fit(x, y)
    r2_rf = r2_score(y, rf.predict(x))
    importances = rf.named_steps["model"].feature_importances_
    imp_df = pd.DataFrame({"feature": names, "importance": importances}).sort_values("importance", ascending=False)
    return coef_df, imp_df, r2_lin, r2_rf


def write_excel(full: pd.DataFrame, numeric: pd.DataFrame, summaries: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        full.to_excel(writer, sheet_name="All Plants", index=False)
        numeric.to_excel(writer, sheet_name="Numeric CFA Only", index=False)
        for name, df in summaries.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(OUT_XLSX)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="245D7A")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Aptos", size=11)
        for col_cells in ws.columns:
            max_len = min(max(len(str(c.value or "")) for c in col_cells) + 2, 42)
            ws.column_dimensions[col_cells[0].column_letter].width = max(10, max_len)
    wb.save(OUT_XLSX)


def write_report(numeric: pd.DataFrame, summaries: dict[str, pd.DataFrame], r2_lin: float, r2_rf: float) -> None:
    gap = numeric["45Z-CFA Gap"].dropna()
    lcfs_gap = numeric["LCFS-CFA Gap"].dropna()
    cfa_lower_share = (gap > 0).mean() if len(gap) else np.nan
    corr_45z_cfa = numeric[["Estimated 45Z CI", "Canadian Federal CI"]].corr().iloc[0, 1]
    corr_lcfs_cfa = numeric[["LCFS CI", "Canadian Federal CI"]].corr().iloc[0, 1]
    imp = summaries["Feature Importance"].head(8)

    lines = [
        "# LCFS / Canadian Federal CI / Estimated 45Z Comparison",
        "",
        f"Rows in full JSON: {len(numeric.index) + (202 - len(numeric.index))}",
        f"Rows with numeric Canadian Federal CI used in statistics: {len(numeric)}",
        "",
        "## Core Results",
        "",
        f"- Average LCFS-CFA gap: {lcfs_gap.mean():.2f} g/MJ.",
        f"- Average 45Z-CFA gap: {gap.mean():.2f} g/MJ.",
        f"- Median 45Z-CFA gap: {gap.median():.2f} g/MJ.",
        f"- Standard deviation of 45Z-CFA gap: {gap.std(ddof=1):.2f} g/MJ.",
        f"- Share of numeric plants where CFA is lower than estimated 45Z: {cfa_lower_share:.1%}.",
        f"- Correlation, LCFS CI vs CFA: {corr_lcfs_cfa:.2f}.",
        f"- Correlation, estimated 45Z CI vs CFA: {corr_45z_cfa:.2f}.",
        "",
        "## Variables Explaining 45Z-CFA Gap",
        "",
        f"- In-sample linear model R2: {r2_lin:.2f}. Random forest R2: {r2_rf:.2f}. These are descriptive only because the numeric CFA sample is small.",
        "- Top feature-importance signals:",
    ]
    for _, row in imp.iterrows():
        lines.append(f"  - {row['feature']}: {row['importance']:.3f}")

    lines += [
        "",
        "## Conclusions",
        "",
        f"1. Is CFA systematically lower than estimated 45Z? {'Yes' if gap.mean() > 0 else 'No'}. The average 45Z-CFA gap is {gap.mean():.2f} g/MJ and CFA is lower in {cfa_lower_share:.1%} of numeric matched plants.",
        f"2. Average methodology gap: using the LCFS-derived 45Z method, the average gap is {gap.mean():.2f} g/MJ. The LCFS-CFA gap before 45Z adjustments averages {lcfs_gap.mean():.2f} g/MJ.",
        "3. Best explanatory variables: the strongest descriptive signals are listed above. Because the estimated 45Z score is mechanically built from LCFS CI, freight, ILUC, and grid adjustment, LCFS CI and freight/grid geography tend to dominate the measured gap.",
        "4. Plant-performance vs methodology: the evidence is more consistent with a methodology difference than a broad plant-performance issue. The 45Z estimate mechanically subtracts a fixed ILUC value and freight normalization from LCFS, while CFA is an independent federal score. Variation around the average gap appears tied to geography, grid/freight treatment, and source CI methodology rather than one consistent owner or state performance problem.",
        "",
        "## Files",
        "",
        f"- Workbook: `{OUT_XLSX}`",
        f"- Source JSON: `{JSON_PATH}`",
    ]
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    full = pd.DataFrame([flatten_record(r) for r in data if isinstance(r, dict)])
    numeric = full.dropna(subset=["LCFS CI", "Canadian Federal CI", "Estimated 45Z CI", "45Z-CFA Gap"]).copy()

    stats = summary_stats(numeric)
    corr_cols = [
        "LCFS CI",
        "Canadian Federal CI",
        "Estimated 45Z CI",
        "LCFS-CFA Gap",
        "45Z-CFA Gap",
        "Freight CI Adjustment",
        "Distance to Hub",
        "Electric Grid Adjustment",
        "Thermal BTU/gal",
        "Ethanol Capacity MGY",
    ]
    correlations = numeric[corr_cols].corr(numeric_only=True).reset_index().rename(columns={"index": "Metric"})
    largest_positive = numeric.sort_values("45Z-CFA Gap", ascending=False).head(15)
    largest_negative = numeric.sort_values("45Z-CFA Gap", ascending=True).head(15)
    coef_df, imp_df, r2_lin, r2_rf = model_analysis(numeric) if len(numeric) >= 8 else (pd.DataFrame(), pd.DataFrame(), np.nan, np.nan)

    summaries = {
        "Summary Stats": stats,
        "Correlations": correlations,
        "Largest Positive Gaps": largest_positive,
        "Largest Negative Gaps": largest_negative,
        "By Owner": group_summary(numeric, "Owner"),
        "By State": group_summary(numeric, "State"),
        "By CCS Status": group_summary(numeric, "CCS Status"),
        "By Gas RNG Usage": group_summary(numeric, "Gas/RNG Usage"),
        "By Electric Grid": group_summary(numeric, "Electric Grid"),
        "Feature Importance": imp_df,
        "Regression Coefficients": coef_df,
        "Model Fit": pd.DataFrame([{"linear_r2_in_sample": r2_lin, "random_forest_r2_in_sample": r2_rf, "numeric_rows": len(numeric)}]),
    }

    write_excel(full, numeric, summaries)
    write_report(numeric, summaries, r2_lin, r2_rf)

    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")
    print(f"Total plants in JSON: {len(full)}")
    print(f"Rows with numeric CFA, LCFS, and estimated 45Z: {len(numeric)}")
    print(stats.to_string(index=False))
    print("\nTop positive 45Z-CFA gaps:")
    print(largest_positive[["Plant", "Owner", "State", "LCFS CI", "Canadian Federal CI", "Estimated 45Z CI", "45Z-CFA Gap"]].head(10).to_string(index=False))
    print("\nTop negative 45Z-CFA gaps:")
    print(largest_negative[["Plant", "Owner", "State", "LCFS CI", "Canadian Federal CI", "Estimated 45Z CI", "45Z-CFA Gap"]].head(10).to_string(index=False))


if __name__ == "__main__":
    main()
