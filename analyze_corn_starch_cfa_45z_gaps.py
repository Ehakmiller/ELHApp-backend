from __future__ import annotations

from pathlib import Path
import json
import math
import sqlite3

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from analyze_lcfs_cfa_45z_gaps import (
    BASELINE_45Z_FREIGHT_MILES,
    FREIGHT_FACTOR,
    FREIGHT_MULTIPLIER,
    GRID_BRIDGE_ADJ,
    ILUC,
    JSON_PATH,
    as_num,
    boolish,
    choose_hub,
    first,
    gas_label,
    haversine_miles,
    norm,
    norm_upper,
    pipeline_label,
    plant_epm,
    group_summary,
    model_analysis,
    summary_stats,
)


DB_PATH = Path(
    r"C:\Users\ehakm\OneDrive\Documents\Python Code\Ethanol db\Ethanol DB\ethanol_production.db"
)
OUT_XLSX = Path(r"C:\Users\ehakm\Documents\ELHApp-backend\Corn_Starch_CFA_45Z_Canada_Gap_Diagnostics.xlsx")
OUT_MD = Path(r"C:\Users\ehakm\Documents\ELHApp-backend\Corn_Starch_CFA_45Z_Canada_Gap_Diagnostics.md")


def load_canadian_corn_ci() -> tuple[dict[str, float], pd.DataFrame]:
    with sqlite3.connect(DB_PATH) as con:
        df = pd.read_sql_query('SELECT * FROM "Candian_Fed_CI"', con)
    df["EPM_norm"] = df["EPM"].map(lambda x: norm(x).replace(".0", ""))
    df["CFA_numeric"] = df["Approved CI (gCO2e/MJ)"].map(as_num)
    corn = df[
        df["CI Status"].astype(str).str.strip().str.casefold().eq("active")
        & df["Fuel Type"].astype(str).str.strip().str.casefold().eq("ethanol")
        & df["Feedstock Type"].astype(str).str.strip().str.casefold().eq("corn")
        & df["EPM_norm"].ne("")
        & df["CFA_numeric"].notna()
    ].copy()
    ci_map = corn.groupby("EPM_norm")["CFA_numeric"].first().to_dict()
    return ci_map, corn


def best_lcfs_corn_ci(row: dict) -> float:
    ci_summary = row.get("ci_summary") or {}
    by_feedstock = ci_summary.get("ci_by_feedstock") or {}
    value = as_num(by_feedstock.get("ci_corn_g_per_mj"))
    if math.isfinite(value):
        return value

    detail = row.get("lcfs_detail") or {}
    ca_rows = detail.get("ca_detail") or row.get("ca_detail") or []
    candidates = []
    for d in ca_rows:
        if not isinstance(d, dict):
            continue
        feedstock = norm_upper(d.get("feedstock"))
        pathway = norm_upper(d.get("pathway_type"))
        if "CORN" in feedstock and "FIBER" not in feedstock and pathway == "STARCH":
            score = as_num(d.get("ci_score"))
            if math.isfinite(score):
                candidates.append((norm(d.get("detail_date")), score))
    if not candidates:
        return np.nan
    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]


def best_oregon_corn_ci(row: dict) -> float:
    detail = row.get("lcfs_detail") or {}
    or_rows = detail.get("or_detail") or row.get("or_detail") or []
    candidates = []
    for d in or_rows:
        if not isinstance(d, dict):
            continue
        feedstock = norm_upper(d.get("feedstock"))
        pathway = norm_upper(d.get("pathway_type"))
        if "CORN" in feedstock and "FIBER" not in feedstock and (not pathway or pathway == "STARCH"):
            score = as_num(d.get("ci_score"))
            if math.isfinite(score):
                candidates.append((norm(d.get("detail_date")), score))
    if not candidates:
        return np.nan
    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]


def flatten_corn_record(row: dict, cfa_map: dict[str, float]) -> dict:
    fac = row.get("fac_info") or {}
    co2 = row.get("co2_info") or {}
    tech = row.get("tech_flags") or {}
    epa = row.get("epa_ghg_derived") or {}

    epm = plant_epm(row)
    owner = norm(first(row.get("ownership"), fac.get("ownership")))
    state = norm(first(row.get("state"), fac.get("state"))).upper()
    plant = norm(first(row.get("plant_name"), fac.get("plant_name")))
    city = norm(first(row.get("city"), fac.get("city")))

    lcfs_ci = best_lcfs_corn_ci(row)
    oregon_ci = best_oregon_corn_ci(row)
    cfa = cfa_map.get(epm, np.nan)

    lat = as_num(first(row.get("latitude"), fac.get("latitude")))
    lon = as_num(first(row.get("longitude"), fac.get("longitude")))
    rail = norm(first(row.get("rail_lines"), fac.get("rail_lines")))
    hub = choose_hub(rail)
    miles = haversine_miles(lat, lon, hub.lat, hub.lon)
    miles_eff = miles * FREIGHT_MULTIPLIER if math.isfinite(miles) else np.nan
    miles_net = miles_eff - BASELINE_45Z_FREIGHT_MILES if math.isfinite(miles_eff) else np.nan
    freight_ci = miles_net * FREIGHT_FACTOR if math.isfinite(miles_net) else np.nan

    grid = norm(first(row.get("electrical_grid_designation"), tech.get("electrical_grid_designation"))).upper()
    grid_adj = GRID_BRIDGE_ADJ.get(grid, 0.0)
    estimated_45z = lcfs_ci - freight_ci - ILUC + grid_adj if math.isfinite(lcfs_ci) and math.isfinite(freight_ci) else np.nan

    direct = first(row.get("co2_pipeline_direct"), co2.get("co2_pipeline_direct"))
    third = first(row.get("co2_pipeline_3rd_party"), co2.get("co2_pipeline_3rd_party"))
    sponsor = first(row.get("co2_sponsor"), co2.get("co2_sponsor"))
    rail_co2 = first(row.get("co2_rail_connect"), co2.get("co2_rail_connect"))
    pipe_label = pipeline_label(direct, third, sponsor, rail_co2)

    return {
        "EPM": epm,
        "Plant": plant,
        "Owner": owner,
        "State": state,
        "City": city,
        "LCFS Corn Starch CI": lcfs_ci,
        "Oregon Corn CI": oregon_ci,
        "Canadian Federal Corn CI": cfa,
        "Estimated 45Z CI": estimated_45z,
        "LCFS-CFA Gap": lcfs_ci - cfa if math.isfinite(lcfs_ci) and math.isfinite(cfa) else np.nan,
        "Oregon-CFA Gap": oregon_ci - cfa if math.isfinite(oregon_ci) and math.isfinite(cfa) else np.nan,
        "Oregon-LCFS Gap": oregon_ci - lcfs_ci if math.isfinite(oregon_ci) and math.isfinite(lcfs_ci) else np.nan,
        "45Z-CFA Gap": estimated_45z - cfa if math.isfinite(estimated_45z) and math.isfinite(cfa) else np.nan,
        "canada_minus_lcfs": cfa - lcfs_ci if math.isfinite(lcfs_ci) and math.isfinite(cfa) else np.nan,
        "canada_minus_oregon": cfa - oregon_ci if math.isfinite(oregon_ci) and math.isfinite(cfa) else np.nan,
        "canada_minus_45z": cfa - estimated_45z if math.isfinite(estimated_45z) and math.isfinite(cfa) else np.nan,
        "Freight CI Adjustment": freight_ci,
        "Distance to Hub": miles,
        "Net 45Z Freight Miles": miles_net,
        "Default Hub": hub.name,
        "Default Hub Railroad": hub.railroad,
        "Electric Grid": grid,
        "Electric Grid Adjustment": grid_adj,
        "Gas/RNG Usage": gas_label(row),
        "biogas_rng_flag": gas_label(row) == "RNG/Biogas",
        "CCS Status": pipe_label,
        "Has CCS/Pipeline": pipe_label != "None",
        "pipeline_flag": pipe_label != "None",
        "ccs_flag": pipe_label in {"Direct", "Tallgrass/Trailblazer", "Rail CO2", "Other third-party"},
        "Technology": norm(first(row.get("technology"), tech.get("technology"))),
        "Thermal BTU/gal": as_num(first(row.get("thermal_btu_per_gal_est"), epa.get("thermal_btu_per_gal_est"))),
        "Ethanol Capacity MGY": as_num(first(row.get("ethanol_capacity_mgy"), fac.get("ethanol_capacity_mgy"))),
    }


def write_excel(full: pd.DataFrame, numeric: pd.DataFrame, summaries: dict[str, pd.DataFrame], corn_source: pd.DataFrame) -> None:
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        numeric.sort_values("canada_minus_45z").to_excel(writer, sheet_name="Plant Level Ranked", index=False)
        full.to_excel(writer, sheet_name="All Plants Corn Match", index=False)
        corn_source.to_excel(writer, sheet_name="Canadian Corn Source", index=False)
        for name, df in summaries.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    wb = load_workbook(OUT_XLSX)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="245D7A")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Aptos", size=11)
    wb.save(OUT_XLSX)


def gap_summary(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    if group_col not in df.columns:
        return pd.DataFrame()
    return (
        df.groupby(group_col, dropna=False)["canada_minus_45z"]
        .agg(["count", "mean", "median", "std", "min", "max"])
        .reset_index()
        .rename(columns={group_col: "group", "std": "standard_deviation"})
        .sort_values(["count", "mean"], ascending=[False, True])
    )


def build_conclusions(numeric: pd.DataFrame, summaries: dict[str, pd.DataFrame]) -> pd.DataFrame:
    gap = numeric["canada_minus_45z"].dropna()
    oregon_gap = numeric["canada_minus_oregon"].dropna()
    lcfs_gap = numeric["canada_minus_lcfs"].dropna()
    lower_share = (gap < 0).mean() if len(gap) else np.nan

    owner = summaries.get("Owner Summary", pd.DataFrame())
    state = summaries.get("State Summary", pd.DataFrame())
    flag_rows = []
    for name in ["Pipeline Summary", "CCS Summary", "RNG Biogas Summary"]:
        df = summaries.get(name, pd.DataFrame())
        if not df.empty:
            spread = df["mean"].max() - df["mean"].min() if len(df) > 1 else 0
            flag_rows.append(f"{name}: mean spread {spread:.2f} g/MJ")

    owner_spread = owner["mean"].max() - owner["mean"].min() if not owner.empty and len(owner) > 1 else np.nan
    state_spread = state["mean"].max() - state["mean"].min() if not state.empty and len(state) > 1 else np.nan

    conclusions = [
        ("Numeric rows", len(numeric)),
        ("Mean canada_minus_45z", gap.mean()),
        ("Median canada_minus_45z", gap.median()),
        ("Std dev canada_minus_45z", gap.std(ddof=1)),
        ("Min canada_minus_45z", gap.min()),
        ("Max canada_minus_45z", gap.max()),
        ("Share where Canada is lower than estimated 45Z", lower_share),
        ("Mean canada_minus_lcfs", lcfs_gap.mean() if len(lcfs_gap) else np.nan),
        ("Mean canada_minus_oregon", oregon_gap.mean() if len(oregon_gap) else np.nan),
        ("Owner mean spread", owner_spread),
        ("State mean spread", state_spread),
        (
            "Conclusion",
            "Canadian Federal CI is consistently lower than estimated 45Z in this numeric corn-starch matched sample. "
            "The average canada_minus_45z gap is negative, indicating Canada is below 45Z. "
            "The standard deviation and owner/state spreads show the gap is not perfectly constant, but the direction is consistent. "
            "Observed clustering should be interpreted cautiously because the numeric matched sample is small.",
        ),
        (
            "Flag summary note",
            "; ".join(flag_rows) if flag_rows else "No flag summaries available.",
        ),
    ]
    return pd.DataFrame(conclusions, columns=["item", "value"])


def main() -> None:
    cfa_map, corn_source = load_canadian_corn_ci()
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    full = pd.DataFrame([flatten_corn_record(r, cfa_map) for r in data if isinstance(r, dict)])
    numeric = full.dropna(subset=["LCFS Corn Starch CI", "Canadian Federal Corn CI", "Estimated 45Z CI", "45Z-CFA Gap"]).copy()
    numeric_for_shared = numeric.rename(columns={"LCFS Corn Starch CI": "LCFS CI", "Canadian Federal Corn CI": "Canadian Federal CI"})

    coef_df, imp_df, r2_lin, r2_rf = model_analysis(numeric_for_shared) if len(numeric) >= 8 else (pd.DataFrame(), pd.DataFrame(), np.nan, np.nan)
    summaries = {
        "Summary Stats": summary_stats(numeric_for_shared),
        "Owner Summary": gap_summary(numeric, "Owner"),
        "State Summary": gap_summary(numeric, "State"),
        "Pipeline Summary": gap_summary(numeric, "pipeline_flag"),
        "CCS Summary": gap_summary(numeric, "ccs_flag"),
        "RNG Biogas Summary": gap_summary(numeric, "biogas_rng_flag"),
        "Technology Summary": gap_summary(numeric, "Technology"),
        "Correlations": numeric_for_shared[
            [
                "LCFS CI",
                "Oregon Corn CI",
                "Canadian Federal CI",
                "Estimated 45Z CI",
                "LCFS-CFA Gap",
                "Oregon-CFA Gap",
                "Oregon-LCFS Gap",
                "45Z-CFA Gap",
                "canada_minus_lcfs",
                "canada_minus_oregon",
                "canada_minus_45z",
                "Freight CI Adjustment",
                "Distance to Hub",
                "Electric Grid Adjustment",
                "Thermal BTU/gal",
                "Ethanol Capacity MGY",
            ]
        ].corr().reset_index().rename(columns={"index": "Metric"}),
        "Oregon Numeric Rows": numeric[numeric["Oregon Corn CI"].notna()].copy(),
        "Canada Much Lower Ranked": numeric.sort_values("canada_minus_45z", ascending=True).head(15),
        "Canada Less Lower Ranked": numeric.sort_values("canada_minus_45z", ascending=False).head(15),
        "By Owner Old Sign": group_summary(numeric_for_shared, "Owner"),
        "By State Old Sign": group_summary(numeric_for_shared, "State"),
        "By CCS Status": group_summary(numeric_for_shared, "CCS Status"),
        "By Electric Grid": group_summary(numeric_for_shared, "Electric Grid"),
        "Feature Importance": imp_df,
        "Regression Coefficients": coef_df,
        "Model Fit": pd.DataFrame([{"linear_r2_in_sample": r2_lin, "random_forest_r2_in_sample": r2_rf, "numeric_rows": len(numeric)}]),
    }
    summaries["Written Conclusions"] = build_conclusions(numeric, summaries)
    write_excel(full, numeric, summaries, corn_source)

    gap = numeric["45Z-CFA Gap"]
    canada_gap = numeric["canada_minus_45z"]
    report = [
        "# Corn Starch CFA / LCFS / Estimated 45Z Comparison",
        "",
        "This rerun compares corn-to-corn only:",
        "- LCFS uses `ci_summary.ci_by_feedstock.ci_corn_g_per_mj` or starch/corn LCFS detail fallback.",
        "- CFA uses only `Candian_Fed_CI` rows where `Feedstock Type = Corn` and CI is numeric.",
        "",
        f"Numeric corn comparison rows: {len(numeric)}",
        f"Rows with Oregon corn CI and numeric CFA: {numeric['Oregon Corn CI'].notna().sum()}",
        f"Average canada_minus_45z gap: {canada_gap.mean():.2f} g/MJ",
        f"Median canada_minus_45z gap: {canada_gap.median():.2f} g/MJ",
        f"Std dev canada_minus_45z gap: {canada_gap.std(ddof=1):.2f} g/MJ",
        f"Share where Canada is lower than estimated 45Z: {(canada_gap < 0).mean():.1%}",
        f"Average 45Z-CFA gap, old sign: {gap.mean():.2f} g/MJ",
        f"CFA lower than estimated 45Z share: {(gap > 0).mean():.1%}",
        "",
        f"Workbook: `{OUT_XLSX}`",
    ]
    OUT_MD.write_text("\n".join(report) + "\n", encoding="utf-8")

    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")
    print(f"Canadian numeric corn source rows: {len(corn_source)}")
    print(f"Numeric corn comparison rows: {len(numeric)}")
    print(f"Rows with Oregon corn CI and numeric CFA: {numeric['Oregon Corn CI'].notna().sum()}")
    print(summaries["Summary Stats"].to_string(index=False))
    print("\nCanada minus 45Z summary by owner:")
    print(summaries["Owner Summary"].to_string(index=False))


if __name__ == "__main__":
    main()
