from pathlib import Path
import sqlite3

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font


DB_PATH = Path(
    r"C:\Users\ehakm\OneDrive\Documents\Python Code\Ethanol db\Ethanol DB\ethanol_production.db"
)
OUT_DIR = Path(r"C:\Users\ehakm\Documents\ELHApp-backend")
OUT_XLSX = OUT_DIR / "Carbon_Sequestration_Plants_By_State.xlsx"
OUT_PNG = OUT_DIR / "Carbon_Sequestration_Plants_By_State.png"


YES_VALUES = {"yes", "y", "true", "1", "direct", "third party", "3rd party"}
NO_VALUES = {"", "no", "n", "false", "0", "none", "nan", "null", "unknown"}
EXCLUDED_THIRD_PARTY_VALUES = {"scs", "summit carbon solutions", "summit"}


def clean_state(value) -> str:
    return str(value or "").replace("\xa0", "").strip().upper()


def has_value(value) -> bool:
    text = str(value or "").replace("\xa0", " ").strip()
    if text.lower() in NO_VALUES:
        return False
    return bool(text) or text.lower() in YES_VALUES


def load_plant_flags(con: sqlite3.Connection) -> pd.DataFrame:
    query = """
    SELECT
      "EPM" AS epm,
      "Name" AS plant_name,
      "State" AS state,
      "C02 Pipeline -Direct" AS direct_ccs,
      "C02 Pipeline -3rd Party" AS third_party_ccs,
      "Sponsor" AS sponsor
    FROM corn_processors
    WHERE TRIM(COALESCE("Status", '')) = 'Active'
      AND TRIM(COALESCE("EPM", '')) <> ''
      AND TRIM(COALESCE("State", '')) <> ''
    """
    df = pd.read_sql_query(query, con)
    df["state"] = df["state"].map(clean_state)
    df["has_direct_ccs"] = df["direct_ccs"].map(has_value)
    df["has_third_party_ccs"] = df["third_party_ccs"].map(has_value)
    df.loc[
        df["third_party_ccs"].fillna("").astype(str).str.strip().str.lower().isin(EXCLUDED_THIRD_PARTY_VALUES),
        "has_third_party_ccs",
    ] = False
    df.loc[
        df["sponsor"].fillna("").astype(str).str.strip().str.lower().isin(EXCLUDED_THIRD_PARTY_VALUES),
        "has_third_party_ccs",
    ] = False
    df["has_any_ccs"] = df["has_direct_ccs"] | df["has_third_party_ccs"]

    plant = (
        df.groupby(["epm", "plant_name", "state"], dropna=False)
        .agg(
            has_direct_ccs=("has_direct_ccs", "max"),
            has_third_party_ccs=("has_third_party_ccs", "max"),
            has_any_ccs=("has_any_ccs", "max"),
            direct_ccs=("direct_ccs", lambda s: first_nonblank(s)),
            third_party_ccs=("third_party_ccs", lambda s: first_nonblank(s)),
            sponsor=("sponsor", lambda s: first_nonblank(s)),
        )
        .reset_index()
    )
    return plant


def first_nonblank(values) -> str:
    for value in values:
        text = str(value or "").replace("\xa0", " ").strip()
        if text and text.lower() not in NO_VALUES:
            return text
    return ""


def summarize_by_state(plant: pd.DataFrame) -> pd.DataFrame:
    summary = (
        plant[plant["has_any_ccs"]]
        .groupby("state", as_index=False)
        .agg(
            plant_count=("epm", "nunique"),
            direct_count=("has_direct_ccs", "sum"),
            third_party_count=("has_third_party_ccs", "sum"),
        )
        .sort_values(["plant_count", "state"], ascending=[False, True])
    )
    return summary


def add_excel_chart(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Aptos", size=11)
                cell.alignment = Alignment(vertical="top")
        ws.freeze_panes = "A2"

    ws = wb["By State"]
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Ethanol Plants with Carbon Sequestration by State"
    chart.y_axis.title = "Plant count"
    chart.x_axis.title = "State"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "F2")
    wb.save(path)


def save_png(summary: pd.DataFrame) -> None:
    plt.rcParams["font.family"] = "Aptos"
    fig, ax = plt.subplots(figsize=(10, 6), constrained_layout=True)
    ax.bar(summary["state"], summary["plant_count"], color="#245D7A")
    ax.set_title(
        "Ethanol Plants with Carbon Sequestration by State",
        fontsize=15,
        fontweight="bold",
        pad=12,
    )
    ax.set_xlabel("")
    ax.set_ylabel("Plant count", fontsize=11)
    ax.grid(axis="y", color="#D9DEE7", linewidth=0.8)
    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color("#AAB2C0")
    ax.spines["bottom"].set_color("#AAB2C0")
    for patch in ax.patches:
        value = patch.get_height()
        ax.annotate(
            f"{value:,.0f}",
            (patch.get_x() + patch.get_width() / 2, value),
            ha="center",
            va="bottom",
            fontsize=9,
            xytext=(0, 3),
            textcoords="offset points",
        )
    fig.savefig(OUT_PNG, dpi=200, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    with sqlite3.connect(DB_PATH) as con:
        plant = load_plant_flags(con)
    summary = summarize_by_state(plant)
    flagged_plants = plant[plant["has_any_ccs"]].sort_values(["state", "plant_name"])

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="By State", index=False)
        flagged_plants.to_excel(writer, sheet_name="Plant Detail", index=False)
    add_excel_chart(OUT_XLSX)
    save_png(summary)

    print(f"Wrote {OUT_PNG}")
    print(f"Wrote {OUT_XLSX}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
