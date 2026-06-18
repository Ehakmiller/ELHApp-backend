from __future__ import annotations

from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
import re
import shutil
import tempfile
from xml.etree import ElementTree as ET
import zipfile

from openpyxl import load_workbook


WORKBOOK = Path(
    r"C:\Users\ehakm\OneDrive\Documents\Ethanol Industry Data\Ethanol Production DataBase\MASTER Plant File - Current.xlsx"
)
WASHINGTON_SHEET = "Washington"
CORN_SHEET = "Corn Processors"
HEADER_ROW = 5

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)

STATE_NAMES = {
    "colorado": "CO",
    "illinois": "IL",
    "indiana": "IN",
    "iowa": "IA",
    "kansas": "KS",
    "michigan": "MI",
    "minnesota": "MN",
    "missouri": "MO",
    "nebraska": "NE",
    "north dakota": "ND",
    "ohio": "OH",
    "oregon": "OR",
    "south dakota": "SD",
    "wisconsin": "WI",
}

GENERIC_PLANT_NAMES = {
    "",
    "any cellulosic biomass including corn stover, wheat straw, or sugarcane straw",
    "any feedstock",
    "corn",
    "midwest corn, dry mill",
    "midwest corn, dry mill, dry dgs",
    "midwest corn, dry mill, modified dgs",
    "sorghum",
    "sugarcane and molasses",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def norm(value: object) -> str:
    text = clean(value).lower()
    text = re.sub(
        r"\b(llc|inc|corp|corporation|company|co|ltd|limited|lp|llp|ethanol)\b",
        " ",
        text,
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_epm(value: object) -> str:
    text = clean(value)
    if not text:
        return ""
    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except ValueError:
        pass
    return text


def split_first_segment(description: str) -> str:
    depth = 0
    for i, char in enumerate(description):
        if char == "(":
            depth += 1
        elif char == ")" and depth:
            depth -= 1
        elif char == ";" and depth == 0:
            return description[:i]
    return description


def clean_plant_name(value: str) -> str:
    text = re.sub(r"\s*\([^)]*\)\s*", " ", clean(value))
    text = re.sub(r"\bEPA\s+(?:Company|Facility)\s+ID\b.*$", "", text, flags=re.I)
    text = re.sub(r"\bFacility\s+Name:\s*", "", text, flags=re.I)
    return clean(text).strip(" ,.-")


def location_from_description(description: str) -> tuple[str, str]:
    text = clean(description)
    patterns = (
        r"produced in\s+([A-Za-z .'-]+?),\s*([A-Z]{2})\b",
        r"produced in\s+([A-Za-z .'-]+?),\s*([A-Za-z ]+?)(?:\s+using|\s+and|;|\.|$)",
        r"produced in\s+([A-Za-z .'-]+?)\s+(North Dakota|South Dakota|Minnesota|Nebraska|Colorado|Kansas|Iowa|Oregon|Illinois|Indiana|Ohio|Wisconsin|Michigan|Missouri)(?:\s+and|;|\.|$)",
    )
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if not match:
            continue
        city = clean(match.group(1)).strip(" ,.-")
        state_raw = clean(match.group(2))
        state = STATE_NAMES.get(state_raw.lower(), state_raw.upper() if len(state_raw) == 2 else "")
        if state:
            return city, state

    state_only = re.search(
        r"produced in\s+(North Dakota|South Dakota|Minnesota|Nebraska|Colorado|Kansas|Iowa|Oregon|Illinois|Indiana|Ohio|Wisconsin|Michigan|Missouri)(?:\s+and|;|\.|$)",
        text,
        flags=re.I,
    )
    if state_only:
        return "", STATE_NAMES[state_only.group(1).lower()]
    return "", ""


def parse_description(description: object) -> tuple[str, str, str]:
    text = clean(description)
    plant = clean_plant_name(split_first_segment(text))
    city, state = location_from_description(text)
    return plant, city, state


def header_map_from_values(values: tuple[object, ...]) -> dict[str, int]:
    return {
        clean(value): index
        for index, value in enumerate(values)
        if clean(value)
    }


def worksheet_header(ws, row_num: int) -> tuple[object, ...]:
    for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
        return row
    raise RuntimeError(f"Could not read header row {row_num} from {ws.title}")


def extract_ids(description: str) -> tuple[set[str], set[str]]:
    text = clean(description)
    company_ids = set()
    facility_ids = set()
    for pattern in (
        r"(?:EPA\s*)?(?:Company\s*ID|CID)\s*[-=]?\s*(\d+)",
        r"\((\d{4});\s*\d{5}\)",
    ):
        company_ids.update(re.findall(pattern, text, flags=re.I))
    for pattern in (
        r"(?:EPA\s*)?(?:Facility\s*ID|FID)\s*[-=]?\s*(\d+)",
        r"\(\d{4};\s*(\d{5})\)",
    ):
        facility_ids.update(re.findall(pattern, text, flags=re.I))
    return company_ids, facility_ids


def load_corn_rows(ws) -> list[dict[str, str]]:
    headers = header_map_from_values(worksheet_header(ws, 1))
    required = ("CA Co ID", "Facility Id", "CA Facility ID", "EPM", "Ownership", "Name", "City", "State")
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f"Missing columns on {CORN_SHEET}: {missing}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[headers["Name"]])
        if not name:
            continue
        city = clean(row[headers["City"]])
        state = clean(row[headers["State"]]).upper().replace(" ", "")
        ownership = clean(row[headers["Ownership"]])
        rows.append(
            {
                "ca_co_id": clean(row[headers["CA Co ID"]]),
                "facility_id": clean(row[headers["Facility Id"]]),
                "ca_facility_id": clean(row[headers["CA Facility ID"]]),
                "epm": normalize_epm(row[headers["EPM"]]),
                "ownership": ownership,
                "name": name,
                "city": city,
                "state": state,
                "name_norm": norm(name),
                "ownership_norm": norm(ownership),
                "city_norm": norm(city),
            }
        )
    return rows


def find_by_ids(
    rows: list[dict[str, str]],
    company_ids: set[str],
    facility_ids: set[str],
) -> tuple[dict[str, str] | None, str]:
    both = []
    for row in rows:
        company_match = row["ca_co_id"] and row["ca_co_id"] in company_ids
        facility_match = row["facility_id"] in facility_ids or row["ca_facility_id"] in facility_ids
        if company_match and facility_match:
            both.append(row)
    if len(both) == 1:
        return both[0], "company+facility id"

    facility_matches = [
        row for row in rows if row["facility_id"] in facility_ids or row["ca_facility_id"] in facility_ids
    ]
    if len(facility_matches) == 1:
        return facility_matches[0], "facility id"

    company_matches = [row for row in rows if row["ca_co_id"] and row["ca_co_id"] in company_ids]
    if len(company_matches) == 1:
        return company_matches[0], "company id"
    return None, ""


def find_by_text(
    rows: list[dict[str, str]],
    description: str,
    plant: str,
    city: str,
    state: str,
) -> tuple[dict[str, str] | None, str, float]:
    candidates = rows
    if state:
        state_candidates = [row for row in candidates if row["state"] == state]
        if state_candidates:
            candidates = state_candidates
    if city:
        city_candidates = [row for row in candidates if row["city_norm"] == norm(city)]
        if city_candidates:
            if len(city_candidates) == 1:
                return city_candidates[0], "unique city/state", 0.99
            candidates = city_candidates

    desc_norm = norm(description)
    plant_norm = norm(plant)
    best = None
    best_score = 0.0
    for row in candidates:
        score = 0.0
        if row["name_norm"] and row["name_norm"] in desc_norm:
            score = max(score, 0.98)
        if row["ownership_norm"] and row["ownership_norm"] in desc_norm:
            score = max(score, 0.96)
        if plant_norm:
            score = max(score, SequenceMatcher(None, plant_norm, row["name_norm"]).ratio())
            score = max(score, SequenceMatcher(None, plant_norm, row["ownership_norm"]).ratio())
        if city and row["city_norm"] == norm(city):
            score += 0.08
        if state and row["state"] == state:
            score += 0.04
        if score > best_score:
            best_score = score
            best = row

    if best and best_score >= 0.78:
        return best, "text", best_score
    return None, "", best_score


def match_corn_row(
    rows: list[dict[str, str]],
    description: str,
    plant: str,
    city: str,
    state: str,
) -> tuple[dict[str, str] | None, str, float]:
    company_ids, facility_ids = extract_ids(description)
    row, method = find_by_ids(rows, company_ids, facility_ids)
    if row:
        return row, method, 1.0
    return find_by_text(rows, description, plant, city, state)


def compute_updates() -> tuple[dict[tuple[int, str], str], list[tuple[int, str, str, str, float]], int, int]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb[WASHINGTON_SHEET]
    corn_ws = wb[CORN_SHEET]
    headers = header_map_from_values(worksheet_header(ws, HEADER_ROW))
    pathway_col = headers["Pathway Description"]
    corn_rows = load_corn_rows(corn_ws)

    updates: dict[tuple[int, str], str] = {
        (HEADER_ROW, "A"): "EPM",
        (HEADER_ROW, "B"): "Plant",
        (HEADER_ROW, "C"): "City",
        (HEADER_ROW, "D"): "State",
    }
    unmatched = []
    pathway_rows = 0
    matched = 0

    for row_num, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True),
        start=HEADER_ROW + 1,
    ):
        description = clean(row[pathway_col] if pathway_col < len(row) else None)
        if not description:
            continue

        pathway_rows += 1
        plant, city, state = parse_description(description)
        corn_row, method, score = match_corn_row(corn_rows, description, plant, city, state)

        epm = ""
        if corn_row:
            matched += 1
            epm = corn_row["epm"]
            if not plant or plant.lower() in GENERIC_PLANT_NAMES:
                plant = corn_row["name"]
            if not city:
                city = corn_row["city"]
            if not state:
                state = corn_row["state"]
        else:
            unmatched.append((row_num, plant, city, state, round(score, 3)))

        updates[(row_num, "A")] = epm
        updates[(row_num, "B")] = plant
        updates[(row_num, "C")] = city
        updates[(row_num, "D")] = state

    wb.close()
    return updates, unmatched, pathway_rows, matched


def sheet_xml_path(xlsx: Path, sheet_name: str) -> str:
    with zipfile.ZipFile(xlsx) as zf:
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    sheet = None
    for candidate in workbook.findall(f".//{{{NS_MAIN}}}sheet"):
        if candidate.attrib.get("name") == sheet_name:
            sheet = candidate
            break
    if sheet is None:
        raise RuntimeError(f"Could not find sheet: {sheet_name}")

    rel_id = sheet.attrib[f"{{{NS_REL}}}id"]
    target = None
    for rel in rels.findall(f".//{{{NS_PKG_REL}}}Relationship"):
        if rel.attrib.get("Id") == rel_id:
            target = rel.attrib["Target"]
            break
    if not target:
        raise RuntimeError(f"Could not find workbook relationship for sheet: {sheet_name}")
    return "xl/" + target.lstrip("/")


def col_number(col_letters: str) -> int:
    number = 0
    for char in col_letters:
        number = number * 26 + ord(char.upper()) - ord("A") + 1
    return number


def cell_sort_key(cell) -> tuple[int, int]:
    ref = cell.attrib["r"]
    match = re.match(r"([A-Z]+)(\d+)", ref)
    if not match:
        return (10**9, 10**9)
    return (int(match.group(2)), col_number(match.group(1)))


def set_inline_string(cell, value: str) -> None:
    for child in list(cell):
        cell.remove(child)
    if value == "":
        cell.attrib.pop("t", None)
        return
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, f"{{{NS_MAIN}}}is")
    text = ET.SubElement(inline, f"{{{NS_MAIN}}}t")
    if value.strip() != value:
        text.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
    text.text = value


def update_sheet_xml(xml_bytes: bytes, updates: dict[tuple[int, str], str]) -> bytes:
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{NS_MAIN}}}sheetData")
    if sheet_data is None:
        raise RuntimeError("Could not find sheetData in Washington worksheet XML")

    rows_by_num = {
        int(row.attrib["r"]): row
        for row in sheet_data.findall(f"{{{NS_MAIN}}}row")
        if row.attrib.get("r", "").isdigit()
    }

    for (row_num, col), value in updates.items():
        row = rows_by_num.get(row_num)
        if row is None:
            row = ET.Element(f"{{{NS_MAIN}}}row", {"r": str(row_num)})
            sheet_data.append(row)
            rows_by_num[row_num] = row

        ref = f"{col}{row_num}"
        cell = None
        for candidate in row.findall(f"{{{NS_MAIN}}}c"):
            if candidate.attrib.get("r") == ref:
                cell = candidate
                break
        if cell is None:
            cell = ET.Element(f"{{{NS_MAIN}}}c", {"r": ref})
            row.append(cell)
        set_inline_string(cell, value)
        row[:] = sorted(row, key=cell_sort_key)

    sheet_data[:] = sorted(
        sheet_data,
        key=lambda row: int(row.attrib.get("r", "999999999")) if row.attrib.get("r", "").isdigit() else 999999999,
    )
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def rewrite_xlsx_sheet(xlsx: Path, xml_path: str, new_xml: bytes) -> None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = Path(tmp.name)

    try:
        with zipfile.ZipFile(xlsx, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == xml_path:
                    zout.writestr(item, new_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))
        shutil.move(str(tmp_path), xlsx)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def main() -> None:
    updates, unmatched, pathway_rows, matched = compute_updates()

    backup = WORKBOOK.with_name(
        f"{WORKBOOK.stem}.backup_washington_{datetime.now():%Y%m%d_%H%M%S}{WORKBOOK.suffix}"
    )
    shutil.copy2(WORKBOOK, backup)

    xml_path = sheet_xml_path(WORKBOOK, WASHINGTON_SHEET)
    with zipfile.ZipFile(WORKBOOK, "r") as zf:
        current_xml = zf.read(xml_path)
    new_xml = update_sheet_xml(current_xml, updates)
    rewrite_xlsx_sheet(WORKBOOK, xml_path, new_xml)

    print(f"Updated workbook: {WORKBOOK}")
    print(f"Backup created: {backup}")
    print(f"Updated worksheet XML: {xml_path}")
    print(f"Washington pathway rows updated: {pathway_rows}")
    print(f"Matched EPM rows: {matched}")
    print(f"Unmatched EPM rows: {len(unmatched)}")
    if unmatched:
        print("Unmatched rows:")
        for row_num, plant, city, state, score in unmatched:
            print(f"  row {row_num}: {plant} | {city} | {state} | best score {score}")


if __name__ == "__main__":
    main()
