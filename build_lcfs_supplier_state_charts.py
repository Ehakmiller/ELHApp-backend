from pathlib import Path
import sqlite3

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font


DB_PATH = Path(
    r"C:\Users\ehakm\OneDrive\Documents\Python Code\Ethanol db\Ethanol DB\ethanol_production.db"
)
OUT_DIR = Path(r"C:\Users\ehakm\Documents\ELHApp-backend")
OUT_XLSX = OUT_DIR / "LCFS_CA_Suppliers_By_State.xlsx"
OUT_PNG = OUT_DIR / "LCFS_CA_Suppliers_By_State.png"
OUT_COUNT_PNG = OUT_DIR / "LCFS_CA_Supplier_Count_By_State.png"
OUT_GALLONS_PNG = OUT_DIR / "LCFS_CA_Gallons_By_State.png"


def table_columns(con: sqlite3.Connection, table: str) -> list[str]:
    return [row[1] for row in con.execute(f'PRAGMA table_info("{table}")').fetchall()]


def find_volume_sources(con: sqlite3.Connection) -> list[tuple[str, list[str]]]:
    terms = ("gallon", "volume", "amount", "quantity", "qty", "gal")
    rows = []
    tables = [
        r[0]
        for r in con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
    ]
    for table in tables:
        hits = [c for c in table_columns(con, table) if any(t in c.lower() for t in terms)]
        if hits:
            rows.append((table, hits))
    return rows


def load_supplier_counts(con: sqlite3.Connection) -> pd.DataFrame:
    query = """
    WITH lcfs AS (
      SELECT DISTINCT CAST(CAST("Facility ID" AS INTEGER) AS TEXT) AS ca_facility_id
      FROM LCFS_Combined_CI
      WHERE "Facility ID" IS NOT NULL
    ),
    corn AS (
      SELECT
        CAST(CAST("CA Facility ID" AS INTEGER) AS TEXT) AS ca_facility_id,
        TRIM(REPLACE("State", CHAR(160), '')) AS state,
        TRIM("Name") AS plant_name
      FROM corn_processors
      WHERE "CA Facility ID" IS NOT NULL
        AND TRIM(REPLACE(COALESCE("State", ''), CHAR(160), '')) <> ''
    )
    SELECT
      corn.state,
      COUNT(DISTINCT lcfs.ca_facility_id) AS supplier_count
    FROM lcfs
    JOIN corn ON corn.ca_facility_id = lcfs.ca_facility_id
    GROUP BY corn.state
    ORDER BY supplier_count DESC, corn.state
    """
    return pd.read_sql_query(query, con)


def load_gallons_by_state(con: sqlite3.Connection) -> pd.DataFrame:
    query = """
    WITH lcfs AS (
      SELECT DISTINCT CAST(CAST("Facility ID" AS INTEGER) AS TEXT) AS ca_facility_id
      FROM LCFS_Combined_CI
      WHERE "Facility ID" IS NOT NULL
    ),
    corn AS (
      SELECT
        CAST(CAST("CA Facility ID" AS INTEGER) AS TEXT) AS ca_facility_id,
        TRIM(REPLACE("State", CHAR(160), '')) AS state,
        MAX(CAST("Ethanol Production" AS REAL)) AS ethanol_gallons,
        MAX(CAST("Ethanol Capacity" AS REAL)) AS capacity_mgy
      FROM corn_processors
      WHERE "CA Facility ID" IS NOT NULL
        AND TRIM(REPLACE(COALESCE("State", ''), CHAR(160), '')) <> ''
      GROUP BY CAST(CAST("CA Facility ID" AS INTEGER) AS TEXT), TRIM(REPLACE("State", CHAR(160), ''))
    )
    SELECT
      corn.state,
      SUM(corn.ethanol_gallons) AS ethanol_gallons,
      SUM(corn.capacity_mgy) AS capacity_mgy
    FROM lcfs
    JOIN corn ON corn.ca_facility_id = lcfs.ca_facility_id
    GROUP BY corn.state
    ORDER BY ethanol_gallons DESC, corn.state
    """
    return pd.read_sql_query(query, con)


def add_excel_charts(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Aptos", size=11)
                cell.alignment = Alignment(vertical="top")
        ws.freeze_panes = "A2"

    count_ws = wb["Supplier Count"]
    count_chart = BarChart()
    count_chart.type = "col"
    count_chart.style = 10
    count_chart.title = "California LCFS Ethanol Suppliers by State"
    count_chart.y_axis.title = "Supplier count"
    count_chart.x_axis.title = "State"
    data = Reference(count_ws, min_col=2, min_row=1, max_row=count_ws.max_row)
    cats = Reference(count_ws, min_col=1, min_row=2, max_row=count_ws.max_row)
    count_chart.add_data(data, titles_from_data=True)
    count_chart.set_categories(cats)
    count_chart.height = 8
    count_chart.width = 15
    count_ws.add_chart(count_chart, "D2")

    gal_ws = wb["Gallons by State"]
    gal_chart = BarChart()
    gal_chart.type = "col"
    gal_chart.style = 10
    gal_chart.title = "LCFS Supplier Ethanol Gallons by State"
    gal_chart.y_axis.title = "Million gallons"
    gal_chart.x_axis.title = "State"
    data = Reference(gal_ws, min_col=2, min_row=1, max_row=gal_ws.max_row)
    cats = Reference(gal_ws, min_col=1, min_row=2, max_row=gal_ws.max_row)
    gal_chart.add_data(data, titles_from_data=True)
    gal_chart.set_categories(cats)
    gal_chart.height = 8
    gal_chart.width = 15
    gal_ws.add_chart(gal_chart, "E2")

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Data notes"
    notes["A1"].font = Font(name="Aptos", size=13, bold=True)
    note_rows = [
        "Supplier count is distinct LCFS_Combined_CI Facility ID joined to corn_processors CA Facility ID.",
        "Gallons by state uses corn_processors Ethanol Production for those LCFS-listed facilities; values are million gallons.",
        "The LCFS_Combined_CI and LCFS_Detail tables do not contain a shipment/transaction gallons column.",
        "A volume-column scan is included so the data source limitation is auditable.",
    ]
    for i, note in enumerate(note_rows, start=2):
        notes[f"A{i}"] = note
        notes[f"A{i}"].font = Font(name="Aptos", size=11)
    notes.column_dimensions["A"].width = 120

    wb.save(path)


def style_axes(ax, title: str, ylabel: str) -> None:
    ax.set_title(title, fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("")
    ax.set_ylabel(ylabel, fontsize=11)
    ax.grid(axis="y", color="#D9DEE7", linewidth=0.8)
    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color("#AAB2C0")
    ax.spines["bottom"].set_color("#AAB2C0")
    ax.tick_params(axis="x", labelrotation=0)


def add_bar_labels(ax, fmt="{:,.0f}") -> None:
    for patch in ax.patches:
        value = patch.get_height()
        if pd.notna(value):
            ax.annotate(
                fmt.format(value),
                (patch.get_x() + patch.get_width() / 2, value),
                ha="center",
                va="bottom",
                fontsize=9,
                xytext=(0, 3),
                textcoords="offset points",
            )


def save_single_chart(df: pd.DataFrame, x_col: str, y_col: str, title: str, ylabel: str, color: str, path: Path) -> None:
    fig, ax = plt.subplots(figsize=(10, 6), constrained_layout=True)
    ax.bar(df[x_col], df[y_col], color=color)
    style_axes(ax, title, ylabel)
    add_bar_labels(ax)
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    if not DB_PATH.exists():
        raise FileNotFoundError(DB_PATH)

    plt.rcParams["font.family"] = "Aptos"
    plt.rcParams["axes.unicode_minus"] = False

    with sqlite3.connect(DB_PATH) as con:
        counts = load_supplier_counts(con)
        gallons = load_gallons_by_state(con)
        volume_sources = find_volume_sources(con)

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        counts.to_excel(writer, sheet_name="Supplier Count", index=False)
        gallons.to_excel(writer, sheet_name="Gallons by State", index=False)
        pd.DataFrame(volume_sources, columns=["table", "volume_like_columns"]).to_excel(
            writer, sheet_name="Volume Column Scan", index=False
        )
    add_excel_charts(OUT_XLSX)

    fig, axes = plt.subplots(1, 2, figsize=(16, 6), constrained_layout=True)
    axes[0].bar(counts["state"], counts["supplier_count"], color="#245D7A")
    style_axes(axes[0], "California LCFS Ethanol Suppliers by State", "Supplier count")
    add_bar_labels(axes[0])

    axes[1].bar(gallons["state"], gallons["ethanol_gallons"], color="#7A4E24")
    style_axes(
        axes[1],
        "LCFS Supplier Ethanol Gallons by State",
        "Ethanol production, million gallons",
    )
    add_bar_labels(axes[1])

    fig.suptitle("California LCFS Ethanol Supply Footprint", fontsize=18, fontweight="bold")
    fig.savefig(OUT_PNG, dpi=200, bbox_inches="tight")
    plt.close(fig)

    save_single_chart(
        counts,
        "state",
        "supplier_count",
        "California LCFS Ethanol Suppliers by State",
        "Supplier count",
        "#245D7A",
        OUT_COUNT_PNG,
    )
    save_single_chart(
        gallons,
        "state",
        "ethanol_gallons",
        "LCFS Supplier Ethanol Gallons by State",
        "Ethanol production, million gallons",
        "#7A4E24",
        OUT_GALLONS_PNG,
    )

    print(f"Wrote {OUT_PNG}")
    print(f"Wrote {OUT_COUNT_PNG}")
    print(f"Wrote {OUT_GALLONS_PNG}")
    print(f"Wrote {OUT_XLSX}")
    if volume_sources:
        print("Volume-like columns found:")
        for table, cols in volume_sources:
            print(f"  {table}: {cols}")
    else:
        print("No gallons/volume-like columns found in this SQLite database.")


if __name__ == "__main__":
    main()
