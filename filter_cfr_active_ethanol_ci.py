from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


SRC = Path(r"C:\Users\ehakm\Downloads\List of Carbon Intensities Under the Clean Fuel Regulations.csv")
OUT = Path(r"C:\Users\ehakm\Documents\ELHApp-backend\CFR_Active_Ethanol_CI.xlsx")


def main() -> None:
    raw = pd.read_csv(SRC, header=None, dtype=str, keep_default_na=False)
    header_idx = raw.index[
        raw.iloc[:, 0].astype(str).str.strip().eq("CI Applicant")
    ][0]

    df = pd.read_csv(SRC, header=header_idx, dtype=str, keep_default_na=False)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df = df.apply(lambda col: col.map(lambda x: str(x).strip()))

    active_ethanol = df[
        df["Fuel Type"].str.casefold().eq("ethanol")
        & df["CI Status"].str.casefold().eq("active")
    ].copy()

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        active_ethanol.to_excel(writer, sheet_name="Active Ethanol CI", index=False)
        df.to_excel(writer, sheet_name="All CFR CI", index=False)
        (
            df.groupby(["Fuel Type", "CI Status"], dropna=False)
            .size()
            .reset_index(name="rows")
            .sort_values(["Fuel Type", "CI Status"])
            .to_excel(writer, sheet_name="Status Summary", index=False)
        )

    wb = load_workbook(OUT)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="245D7A")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Aptos", size=11)
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_len = min(
                max(len(str(cell.value or "")) for cell in col_cells) + 2,
                55,
            )
            ws.column_dimensions[letter].width = max(12, max_len)
    wb.save(OUT)

    print(f"Wrote {OUT}")
    print(f"All rows: {len(df)}")
    print(f"Active ethanol rows: {len(active_ethanol)}")
    print(active_ethanol[["CI Applicant", "Facility Name", "Facility Location (City/Municipality, Country)", "Approved CI (gCO2e/MJ)", "CI Status"]].to_string(index=False))


if __name__ == "__main__":
    main()
